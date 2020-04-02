import xml.etree.ElementTree as ET

from xml.dom.minidom import Document
import openpyxl
from CZModule.writexml import writexml


def ExcelToxml():
  try:
          wkbook=openpyxl.load_workbook("F:/CUJS_changzhou_LTE_CR_20200326.xlsx")
          wsheet=wkbook.get_sheet_by_name("sheet5")
          nrow=wsheet.max_row
          doc=Document()
          root=doc.createElement("SubnetWork")
          doc.appendChild(root)
          if nrow>=2:
             tempdic={}
             for i in range(2,nrow+1):
                 EnodebID=str(wsheet.cell(i,1).value).strip()
                 MO=str(wsheet.cell(i,3).value).strip()
                 parameter=str(wsheet.cell(i,4).value).strip()
                 value=str(wsheet.cell(i,5).value).strip()


                 ERBS=doc.createElement("ERBS")
                 ERBS.setAttribute("name",EnodebID)
                 root.appendChild(ERBS)

                 MOName=doc.createElement("MO")
                 MOName.setAttribute("MOname",MO)
                 ERBS.appendChild(MOName)

                 paramater=doc.createElement("paramater")
                 paramater_text=doc.createTextNode(value)
                 paramater.appendChild(paramater_text)
                 MOName.appendChild(paramater)

          with open("F:/202004011111.xml",'w') as xmlf:
               doc.writexml(xmlf,addindent=' ',encoding='utf-8',newl="\n")
  finally:
          wkbook.close()


def Excelxml2():
    try:
        wkbook = openpyxl.load_workbook("F:/CUJS_changzhou_LTE_CR_20200326.xlsx")
        wsheet = wkbook.get_sheet_by_name("sheet5")
        nrow = wsheet.max_row
        xml=writexml()
        root=xml.create_node("SubnetWork")
        xml.doc.appendChild(root)
        if nrow>=2:
           tempdic={}
           MOdic={}
           for i in range(2, nrow + 1):
               EnodebID = str(wsheet.cell(i, 1).value).strip()
               MO=str(wsheet.cell(i, 3).value).strip()
               parameter = str(wsheet.cell(i, 4).value).strip()
               value = str(wsheet.cell(i, 5).value).strip()
               if  EnodebID not in tempdic.keys():
                   ERBS = xml.create_node("ERBS", attribute_name="name", attribute_value=EnodebID)
                   root.appendChild(ERBS)
                   tempdic[EnodebID]=ERBS
               if (EnodebID+MO) not in MOdic.keys():
                   MOname = xml.create_node("MO", attribute_name="name", attribute_value=MO)
                   tempdic[EnodebID].appendChild(MOname)
                   MOdic[EnodebID+MO]=MOname


               paramater=xml.create_node(parameter,nodetext=value)
               MOdic[EnodebID+MO].appendChild(paramater)

        xml.savexml("F:/2020040221111.xml")

    finally:

          wkbook.close()


Excelxml2()



def www():
   f=open("F:/CZEZFL3414.xml")
   et=ET.parse(f)
   root=et.getroot()

   for mi in root.findall(".//mi"):
       mtL=[]
       rl=[]
       MO=[]
       for mt in mi.findall("mt"):
           mtL.append(mt.text)

       for r in mi.findall(".//r"):
           rl.append(r.text)

       for o in  mi.findall(".//moid"):
           MO.append(o.text)

       print(mtL,MO,rl)



def modify():
    f=open("F:/CZEZFL3414.xml")
    et=ET.parse(f)
    root=et.getroot()


    print(root)
    a="pmActiveDrbDlSum"


    mi=root.find('.//mi[mt="%s"]/.'%a)
    print(mi)

    for i,j in enumerate(mi.findall("mt")):
        if j.text==a:
           index=i
           break

    for l in mi.findall("mv"):
        l[index+1].text="80000000"


    et.write("F://20000000.xml")







#for child in root.findall(".//mi"):
#    index=0
#    for i in child.findall(".//mt"):
#        index+=1
#        if i.text=="pmSctpStatRetransChunks":
#           child.find("mv")[index].text="1111111111"
#           break
#
#
#et.write("F://12333333.xml")








