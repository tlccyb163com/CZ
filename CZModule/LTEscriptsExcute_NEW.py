#!C:\Users\tlccyb1\AppData\Local\Programs\Python\Python36-32\python
# -*- coding: utf8 -*-

import os,xlrd,xlwt
import sys
import time
import paramiko
import openpyxl
from CZModule import lteericssondt


class ltescriptsExcute(object):

     def __init__(self,workbook):
         self.workbook=workbook

     def paramater(self):
         WK=openpyxl.load_workbook(self.workbook)
         Wsheet=WK.get_sheet_by_name("paramater")
         nrow=Wsheet.max_row
         if nrow>=2:
            dict={}
            output=os.path.abspath(os.path.join(os.getcwd(),"output_paramater_%s"%time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                ERBS=str(Wsheet.cell(row,1).value).strip()
                action=str(Wsheet.cell(row,2).value).strip()
                mo=str(Wsheet.cell(row,3).value).strip()
                parameter=str(Wsheet.cell(row,4).value).strip()
                value=str(Wsheet.cell(row,5).value).strip()
                f=open(os.path.abspath(os.path.join(output,"%s.mos"%ERBS)),'a')
                f.write(lteericssondt.paramater(action=action,mo=mo,canshu=parameter,value=value))
                f.close()
     def DUS52_3Grelation(self):
         WK = openpyxl.load_workbook(self.workbook)
         Wsheet = WK.get_sheet_by_name("DUS52_3Grelation")
         nrow = Wsheet.max_row
         if nrow>=2:
            dict={}
            output=os.path.abspath(os.path.join(os.getcwd(),"output_DUS52_3Grelation_%s"%time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                ERBS=str(Wsheet.cell(row,1).value).strip()
                eNBId=str(Wsheet.cell(row,2).value).strip()
                EutranCellFDD=str(Wsheet.cell(row,3).value).strip()
                CellID=str(Wsheet.cell(row,4).value).strip()
                WCDMACELLNAME=str(Wsheet.cell(row,5).value).strip()
                RNCid=str(Wsheet.cell(row,6).value).strip()
                LAC=str(Wsheet.cell(row,7).value).strip()
                RAC=str(Wsheet.cell(row,8).value).strip()
                rimCapable=str(Wsheet.cell(row,9).value).strip()
                coverageIndicator=str(Wsheet.cell(row,10).value).strip()
                WCellID=str(Wsheet.cell(row,11).value).strip()
                mcc=str(Wsheet.cell(row,12).value).strip()
                mnc=str(Wsheet.cell(row,13).value).strip()
                mncLength=str(Wsheet.cell(row,14).value).strip()
                DLfrequency=str(Wsheet.cell(row,15).value).strip()
                physicalCellIdentity=str(Wsheet.cell(row,16).value).strip()
                isHOallowed=str(Wsheet.cell(row,17).value).strip()
                cellReselectionPriority=str(Wsheet.cell(row,18).value).strip()
                connectedModeMobilityPrio=str(Wsheet.cell(row,19).value).strip()
                csFallbackPrio=str(Wsheet.cell(row,20).value).strip()
                csFallbackPrioEC=str(Wsheet.cell(row,21).value).strip()
                mobilityActionCsfb=str(Wsheet.cell(row,22).value).strip()
                f = open(os.path.abspath(os.path.join(output, "%s.mos" % ERBS)), 'a')
                if not (ERBS+"UtraNetwork","UtraNetwork") in dict.items():
                   f.write(lteericssondt.crUtraNetwork())
                   dict[ERBS+"UtraNetwork"]="UtraNetwork"
                if not (ERBS+DLfrequency,DLfrequency) in dict.items():
                   f.write(lteericssondt.crUtranFrequency(DLfrequency=DLfrequency))
                   dict[ERBS+DLfrequency]=DLfrequency
                if not (EutranCellFDD+DLfrequency,DLfrequency) in dict.items():
                    f.write(lteericssondt.crUtranFreqRelation(EUtranCellFDD=EutranCellFDD,UtranFreqRelation=DLfrequency,cellReselectionPriority=cellReselectionPriority,connectedModeMobilityPrio=connectedModeMobilityPrio,csFallbackPrio=csFallbackPrio,csFallbackPrioEC=csFallbackPrioEC,mobilityActionCsfb=mobilityActionCsfb))
                    dict[EutranCellFDD+DLfrequency]=DLfrequency
                if not (ERBS+WCDMACELLNAME,WCDMACELLNAME) in dict.items():
                    f.write(lteericssondt.crExternalUtranCellFDD(DLfrequency=DLfrequency,RNCid=RNCid,cId=WCellID,physicalCellIdentity=physicalCellIdentity,mcc=mcc,mnc=mnc,mncLength=mncLength,rac=RAC,lac=LAC))
                    dict[ERBS+WCDMACELLNAME]=WCDMACELLNAME
                f.write(lteericssondt.crUtranCellRelation(EUtranCellFDD=EutranCellFDD,DLfrequency=DLfrequency,cId=WCellID,RNCid=RNCid))
                f.close()

     def DUS52_4Grelation(self):
         WK = openpyxl.load_workbook(self.workbook)
         Wsheet = WK.get_sheet_by_name("DUS52_4Grelation")
         nrow = Wsheet.max_row
         if nrow >= 2:
            dict={}
            output = os.path.abspath(os.path.join(os.getcwd(),"output_DUS52_4Grelation_%s" % time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                ERBS=str(Wsheet.cell(row,1).value).strip()
                EUtranCellFDD=str(Wsheet.cell(row,2).value).strip()
                EUtranFreqRelation=str(Wsheet.cell(row,3).value).strip()
                cellReselectionPriority=str(Wsheet.cell(row,4).value).strip()
                connectedModeMobilityPrio=str(Wsheet.cell(row,5).value).strip()
                ExternalENodeBFunction=str(Wsheet.cell(row,6).value).strip()
                ENodeBId=str(Wsheet.cell(row,7).value).strip()
                EUtranCellRelation=str(Wsheet.cell(row,8).value).strip()
                mcc=str(Wsheet.cell(row,9).value).strip()
                mnc=str(Wsheet.cell(row,10).value).strip()
                mncLength=str(Wsheet.cell(row,11).value).strip()
                CellId=str(Wsheet.cell(row,12).value).strip()
                physicalLayerCellIdGroup=str(Wsheet.cell(row,13).value).strip()
                physicalLayerSubCellId=str(Wsheet.cell(row,14).value).strip()
                tac=str(Wsheet.cell(row,15).value).strip()
                ipAddress=str(Wsheet.cell(row,16).value).strip()
                f=open(os.path.abspath(os.path.join(output, "%s.mos" % ERBS)), 'a')
                if ERBS==EUtranCellRelation.split("_")[0]:
                   f.write(lteericssondt.crinsideEUtranCellRelation(EUtranCellFDD=EUtranCellFDD,EUtranFreqRelation=EUtranFreqRelation,EUtranCellRelation=EUtranCellRelation))
                else:
                    if not (ERBS+EUtranFreqRelation,EUtranFreqRelation) in dict.items():
                            f.write(lteericssondt.crEUtranFrequency(EUtranFreqRelation=EUtranFreqRelation))
                            dict[ERBS+EUtranFreqRelation]=EUtranFreqRelation
                    if not (EUtranCellFDD+EUtranFreqRelation,EUtranFreqRelation) in dict.items():
                            f.write(lteericssondt.crEUtranFreqRelation(EUtranCellFDD=EUtranCellFDD,EUtranFreqRelation=EUtranFreqRelation,cellReselectionPriority=cellReselectionPriority,connectedModeMobilityPrio=connectedModeMobilityPrio))
                            dict[EUtranCellFDD+EUtranFreqRelation]=EUtranFreqRelation
                    if not (ERBS+ExternalENodeBFunction,ExternalENodeBFunction) in dict.items():
                            f.write(lteericssondt.crExternalENodeBFunction(ExternalENodeBFunction=ExternalENodeBFunction,mcc=mcc,mnc=mnc,mncLength=mncLength,ENodeBId=ENodeBId))
                            dict[ERBS+ExternalENodeBFunction]=ExternalENodeBFunction
                    if not (ERBS+ExternalENodeBFunction+ExternalENodeBFunction,ExternalENodeBFunction) in dict.items():
                            f.write(lteericssondt.crTermPointToENB(ExternalENodeBFunction=ExternalENodeBFunction,ipAddress=ipAddress))
                            dict[ERBS+ExternalENodeBFunction+ExternalENodeBFunction]=ExternalENodeBFunction
                    if not (ERBS+EUtranCellRelation,EUtranCellRelation) in dict.items():
                            f.write(lteericssondt.crExternalEUtranCellFDD(ExternalENodeBFunction=ExternalENodeBFunction,EUtranCellRelation=EUtranCellRelation,CellId=CellId,physicalLayerCellIdGroup=physicalLayerCellIdGroup,physicalLayerSubCellId=physicalLayerSubCellId,tac=tac,EUtranFreqRelation=EUtranFreqRelation))
                            dict[ERBS+EUtranCellRelation]=EUtranCellRelation
                    f.write(lteericssondt.croutEUtranCellRelation(EUtranCellFDD=EUtranCellFDD,EUtranFreqRelation=EUtranFreqRelation,EUtranCellRelation=EUtranCellRelation,ExternalENodeBFunction=ExternalENodeBFunction))
     def DUS52_2Grelation(self):
         WK = openpyxl.load_workbook(self.workbook)
         Wsheet = WK.get_sheet_by_name("DUS52_2Grelation")
         nrow = Wsheet.max_row
         if nrow>=2:
            dict={}
            output = os.path.abspath(os.path.join(os.getcwd(), "output_DUS52_2Grelation_%s" % time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                ERBS=str(Wsheet.cell(row,1).value).strip()
                eNBId=str(Wsheet.cell(row,2).value).strip()
                EutranCellFDD=str(Wsheet.cell(row,3).value).strip()
                CellID=str(Wsheet.cell(row,4).value).strip()
                GSMNAME=str(Wsheet.cell(row,5).value).strip()
                BSCid=str(Wsheet.cell(row,6).value).strip()
                GeranFreqGroup=str(Wsheet.cell(row,7).value).strip()
                mcc=str(Wsheet.cell(row,8).value).strip()
                mnc=str(Wsheet.cell(row,9).value).strip()
                mncLength=str(Wsheet.cell(row,10).value).strip()
                lac=str(Wsheet.cell(row,11).value).strip()
                ci=str(Wsheet.cell(row,12).value).strip()
                bcch=str(Wsheet.cell(row,13).value).strip()
                ncc=str(Wsheet.cell(row,14).value).strip()
                bcc=str(Wsheet.cell(row,15).value).strip()
                csFallbackPrio=str(Wsheet.cell(row,16).value).strip()
                connectedModeMobilityPrio=str(Wsheet.cell(row,17).value).strip()
                bandIndicator=str(Wsheet.cell(row,18).value).strip()
                rimCapable=str(Wsheet.cell(row,19).value).strip()
                coverageIndicator=str(Wsheet.cell(row,20).value).strip()
                csFallbackPrioEC=str(Wsheet.cell(row,21).value).strip()
                threshXLow=str(Wsheet.cell(row,22).value).strip()
                qRxLevMin=str(Wsheet.cell(row,23).value).strip()
                mobilityAction=str(Wsheet.cell(row,24).value).strip()
                mobilityActionCsfb=str(Wsheet.cell(row,25).value).strip()
                cellReselectionPriority=str(Wsheet.cell(row,26).value).strip()
                f = open(os.path.abspath(os.path.join(output, "%s.mos" % ERBS)), 'a')
                if not (ERBS+"GeraNetwork","GeraNetwork") in dict.items():
                        f.write(lteericssondt.crGeraNetwork())
                        dict[ERBS+"GeraNetwork"]="GeraNetwork"
                if not (ERBS+GeranFreqGroup,GeranFreqGroup) in dict.items():
                        f.write(lteericssondt.crGeranFreqGroup(GeranFreqGroup))
                        dict[ERBS+GeranFreqGroup,GeranFreqGroup]=GeranFreqGroup
                if not (EutranCellFDD+GeranFreqGroup,GeranFreqGroup) in dict.items():
                        f.write(lteericssondt.crGeranFreqGroupRelation(EUtranCellFDD=EutranCellFDD,GeranFreqGroup=GeranFreqGroup,cellReselectionPriority=cellReselectionPriority,connectedModeMobilityPrio=connectedModeMobilityPrio,csFallbackPrio=csFallbackPrio,csFallbackPrioEC=csFallbackPrioEC,threshXLow=threshXLow,qRxLevMin=qRxLevMin,mobilityAction=mobilityAction,mobilityActionCsfb=mobilityActionCsfb))
                        dict[EutranCellFDD+GeranFreqGroup,GeranFreqGroup]=GeranFreqGroup
                if not (ERBS+GeranFreqGroup+bcch,bcch) in dict.items():
                        f.write(lteericssondt.crGeranFrequency(GeranFreqGroup=GeranFreqGroup,BCCH=bcch))
                        dict[ERBS+GeranFreqGroup+bcch]=bcch
                if not (ERBS+GeranFreqGroup+GSMNAME,GSMNAME) in dict.items():
                        f.write(lteericssondt.crExternalGeranCell(GeranFreqGroup=GeranFreqGroup,BCCH=bcch,mcc=mcc,mnc=mnc,mncLength=mncLength,LAC=lac,CI=ci,BCC=bcc,NCC=ncc,GSMCELLNAME=GSMNAME,rimCapable=rimCapable))
                        dict[ERBS+GeranFreqGroup+GSMNAME]=GSMNAME
                f.write(lteericssondt.crGeranCellRelation(EUtranCellFDD=EutranCellFDD,GeranFreqGroup=GeranFreqGroup,GSMCELLNAME=GSMNAME,coverageIndicator=coverageIndicator))



     def MeasBasedCsfbTargetSelection(self):
         WK = openpyxl.load_workbook(self.workbook)
         Wsheet = WK.get_sheet_by_name("MeasBasedCsfbTargetSelection")
         nrow = Wsheet.max_row
         if nrow>=2:
            dict={}
            output = os.path.abspath(os.path.join(os.getcwd(), "output_MeasBasedCsfbTargetSelection_%s" % time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                ERBS=str(Wsheet.cell(row,1).value).strip()
                eNBId=str(Wsheet.cell(row,2).value).strip()
                EutranCellFDD=str(Wsheet.cell(row,3).value).strip()
                CellID=str(Wsheet.cell(row,4).value).strip()
                GSMNAME=str(Wsheet.cell(row,5).value).strip()
                BSCid=str(Wsheet.cell(row,6).value).strip()
                GeranFreqGroup=str(Wsheet.cell(row,7).value).strip()
                mcc=str(Wsheet.cell(row,8).value).strip()
                mnc=str(Wsheet.cell(row,9).value).strip()
                mncLength=str(Wsheet.cell(row,10).value).strip()
                lac=str(Wsheet.cell(row,11).value).strip()
                ci=str(Wsheet.cell(row,12).value).strip()
                bcch=str(Wsheet.cell(row,13).value).strip()
                ncc=str(Wsheet.cell(row,14).value).strip()
                bcc=str(Wsheet.cell(row,15).value).strip()
                csFallbackPrio=str(Wsheet.cell(row,16).value).strip()
                connectedModeMobilityPrio=str(Wsheet.cell(row,17).value).strip()
                bandIndicator=str(Wsheet.cell(row,18).value).strip()
                rimCapable=str(Wsheet.cell(row,19).value).strip()
                coverageIndicator=str(Wsheet.cell(row,20).value).strip()
                csFallbackPrioEC=str(Wsheet.cell(row,21).value).strip()
                threshXLow=str(Wsheet.cell(row,22).value).strip()
                qRxLevMin=str(Wsheet.cell(row,23).value).strip()
                mobilityAction=str(Wsheet.cell(row,24).value).strip()
                mobilityActionCsfb=str(Wsheet.cell(row,25).value).strip()
                cellReselectionPriority=str(Wsheet.cell(row,26).value).strip()
                thresholdRscp=str(Wsheet.cell(row,27).value).strip()
                altCsfbTargetPrio=str(Wsheet.cell(row,27).value).strip()
                f = open(os.path.abspath(os.path.join(output, "%s.mos" % ERBS)), 'a')
                if not (ERBS+"GeraNetwork","GeraNetwork") in dict.items():
                        f.write(lteericssondt.crGeraNetwork())
                        dict[ERBS+"GeraNetwork"]="GeraNetwork"
                if not (ERBS+GeranFreqGroup,GeranFreqGroup) in dict.items():
                        f.write(lteericssondt.crGeranFreqGroup(GeranFreqGroup))
                        dict[ERBS+GeranFreqGroup,GeranFreqGroup]=GeranFreqGroup
                if not (EutranCellFDD+GeranFreqGroup,GeranFreqGroup) in dict.items():
                        f.write(lteericssondt.crGeranFreqGroupRelation(EUtranCellFDD=EutranCellFDD,GeranFreqGroup=GeranFreqGroup,cellReselectionPriority=cellReselectionPriority,connectedModeMobilityPrio=connectedModeMobilityPrio,csFallbackPrio=csFallbackPrio,csFallbackPrioEC=csFallbackPrioEC,threshXLow=threshXLow,qRxLevMin=qRxLevMin,mobilityAction=mobilityAction,mobilityActionCsfb=mobilityActionCsfb))
                        dict[EutranCellFDD+GeranFreqGroup,GeranFreqGroup]=GeranFreqGroup
                if not (ERBS+GeranFreqGroup+bcch,bcch) in dict.items():
                        f.write(lteericssondt.crGeranFrequency(GeranFreqGroup=GeranFreqGroup,BCCH=bcch))
                        dict[ERBS+GeranFreqGroup+bcch]=bcch
                if not (ERBS+GeranFreqGroup+GSMNAME,GSMNAME) in dict.items():
                        f.write(lteericssondt.crExternalGeranCell(GeranFreqGroup=GeranFreqGroup,BCCH=bcch,mcc=mcc,mnc=mnc,mncLength=mncLength,LAC=lac,CI=ci,BCC=bcc,NCC=ncc,GSMCELLNAME=GSMNAME,rimCapable=rimCapable))
                        dict[ERBS+GeranFreqGroup+GSMNAME]=GSMNAME
                f.write(lteericssondt.crGeranCellRelation(EUtranCellFDD=EutranCellFDD,GeranFreqGroup=GeranFreqGroup,GSMCELLNAME=GSMNAME,coverageIndicator=coverageIndicator))

                if not (EutranCellFDD+thresholdRscp,altCsfbTargetPrio) in dict.items():
                        f.write(lteericssondt.crMeasBasedCsfbTargetSelection(EutranCellFDD=EutranCellFDD,thresholdRscp=thresholdRscp,GeranFreqGroupRelation=GeranFreqGroup,altCsfbTargetPrio=altCsfbTargetPrio))
                        dict[EutranCellFDD+thresholdRscp]=altCsfbTargetPrio
                f.close()


     def EUtranFreqRelation(self):
         WK = openpyxl.load_workbook(self.workbook)
         Wsheet = WK.get_sheet_by_name("EUtranFreqRelation")
         nrow = Wsheet.max_row
         if nrow>=2:
             dict = {}
             output = os.path.abspath(os.path.join(os.getcwd(),"output_EUtranFreqRelation_%s" % time.strftime("%y%m%d%H%M%S", time.localtime())))
             os.mkdir(output)
             for row in range(2,nrow+1):
                 ERBS=str(Wsheet.cell(row,1).value).strip()
                 EUtranCellFDD=str(Wsheet.cell(row,2).value).strip()
                 EUtranFreqRelation=str(Wsheet.cell(row,3).value).strip()
                 cellReselectionPriority=str(Wsheet.cell(row,4).value).strip()
                 connectedModeMobilityPrio=str(Wsheet.cell(row,5).value).strip()
                 f = open(os.path.abspath(os.path.join(output, "%s.mos" % ERBS)), 'a')
                 if not (ERBS+EUtranFreqRelation,EUtranFreqRelation) in dict.items():
                         f.write(lteericssondt.crEUtranFrequency(EUtranFreqRelation=EUtranFreqRelation))
                         dict[ERBS+EUtranFreqRelation]=EUtranFreqRelation
                 f.write(lteericssondt.crEUtranFreqRelation(EUtranCellFDD=EUtranCellFDD,EUtranFreqRelation=EUtranFreqRelation,cellReselectionPriority=cellReselectionPriority,connectedModeMobilityPrio=connectedModeMobilityPrio))
                 f.close()
                 
                 
     def changeFrequency(self):
         WK=openpyxl.load_workbook(self.workbook)
         Wsheet=WK.get_sheet_by_name("changeFrequency")
         nrow=Wsheet.max_row
         if nrow>=2:
            dict={}
            output=os.path.abspath(os.path.join(os.getcwd(),"output_changeFrequency_%s"%time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                ERBS=str(Wsheet.cell(row,1).value).strip()
                action=str(Wsheet.cell(row,2).value).strip()
                ExternalUtranCellFDD=str(Wsheet.cell(row,3).value).strip()
                parameter=str(Wsheet.cell(row,4).value).strip()
                UtranFrequency=str(Wsheet.cell(row,5).value).strip()
                with open(os.path.abspath(os.path.join(output,"%s.mos"%ERBS)),'a') as f:
                     f.write(lteericssondt.changeFrequency(ERBS=ERBS,action=action,ExternalUtranCellFDD=ExternalUtranCellFDD,parameter=parameter,UtranFrequency=UtranFrequency))                

     def utranrelation(self):
         WK=openpyxl.load_workbook(self.workbook)
         Wsheet=WK.get_sheet_by_name("utranrelation")
         nrow=Wsheet.max_row
         if nrow>=2:
            dict={}
            output=os.path.abspath(os.path.join(os.getcwd(),"output_utranrelation_%s"%time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                SourceRNC=str(Wsheet.cell(row,1).value).strip()
                SourceCell=str(Wsheet.cell(row,2).value).strip()
                TargetRNC=str(Wsheet.cell(row,3).value).strip()
                TargetCell=str(Wsheet.cell(row,4).value).strip()
                hcsPrio="0" if Wsheet.cell(row,5).value is None else str(Wsheet.cell(row,5).value).strip()
                qHcs="0" if Wsheet.cell(row,6).value is None else str(Wsheet.cell(row,6).value).strip()
                penaltyTime="0" if Wsheet.cell(row,7).value is None else str(Wsheet.cell(row,7).value).strip()
                temporaryOffset1="0" if Wsheet.cell(row,8).value is None else str(Wsheet.cell(row,8).value).strip()
                temporaryOffset2="0" if Wsheet.cell(row,9).value is None else str(Wsheet.cell(row,9).value).strip()
                loadSharingCandidate="0" if Wsheet.cell(row,10).value is None else str(Wsheet.cell(row,10).value).strip()
                qOffset1sn="0" if Wsheet.cell(row,11).value is None else str(Wsheet.cell(row,11).value).strip()
                qOffset2sn="0" if Wsheet.cell(row,12).value is None else str(Wsheet.cell(row,12).value).strip()
                selectionPriority="0" if Wsheet.cell(row,13).value is None else str(Wsheet.cell(row,13).value).strip()
                with open(os.path.abspath(os.path.join(output,"%s.mo"%SourceRNC)),'a') as f:
                     f.write(lteericssondt.crutranrelation(SourceRNC=SourceRNC,SourceCell=SourceCell,TargetRNC=TargetRNC,TargetCell=TargetCell,hcsPrio=hcsPrio,qHcs=qHcs,penaltyTime=penaltyTime,temporaryOffset1=temporaryOffset1,temporaryOffset2=temporaryOffset2,loadSharingCandidate=loadSharingCandidate,qOffset1sn=qOffset1sn,qOffset2sn=qOffset2sn,selectionPriority=selectionPriority))                


     def gsmneighbor(self):
         WK=openpyxl.load_workbook(self.workbook)
         Wsheet=WK.get_sheet_by_name("gsmneighbor")
         nrow=Wsheet.max_row
         if nrow>=2:
            dict={}
            output=os.path.abspath(os.path.join(os.getcwd(),"output_gsmneighbor_%s"%time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                SourceRNC=str(Wsheet.cell(row,1).value).strip()
                SourceCell=str(Wsheet.cell(row,2).value).strip()
                TargetGsmNetwork=str(Wsheet.cell(row,3).value).strip()
                TargetCell=str(Wsheet.cell(row,4).value).strip()
                mobilityRelationType=str(Wsheet.cell(row,5).value).strip()
                qOffset1sn=str(Wsheet.cell(row,6).value).strip()
                selectionPriority=str(Wsheet.cell(row,7).value).strip()
                with open(os.path.abspath(os.path.join(output,"%s.mo"%SourceRNC)),'a') as f:
                     f.write(lteericssondt.crgsmneighbor(SourceRNC=SourceRNC,SourceCell=SourceCell,TargetGsmNetwork=TargetGsmNetwork,TargetCell=TargetCell,mobilityRelationType=mobilityRelationType,qOffset1sn=qOffset1sn,selectionPriority=selectionPriority))    


     def externalutrancell(self):
         WK=openpyxl.load_workbook(self.workbook)
         Wsheet=WK.get_sheet_by_name("externalutrancell")
         nrow=Wsheet.max_row
         if nrow>=2:
            dict={}
            output=os.path.abspath(os.path.join(os.getcwd(),"output_externalutrancell_%s"%time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                RNCName=str(Wsheet.cell(row,1).value).strip()
                ExternalCell=str(Wsheet.cell(row,2).value).strip()
                IurLinkID=str(Wsheet.cell(row,3).value).strip()
                CID=str(Wsheet.cell(row,4).value).strip()
                LAC=str(Wsheet.cell(row,5).value).strip()
                RAC=str(Wsheet.cell(row,6).value).strip()
                PSC=str(Wsheet.cell(row,7).value).strip()
                uarfcnDl=str(Wsheet.cell(row,8).value).strip()
                uarfcnUl=str(Wsheet.cell(row,9).value).strip()
                agpsEnabled=str(Wsheet.cell(row,10).value).strip()
                individualOffset=str(Wsheet.cell(row,11).value).strip()
                maxTxPowerUl=str(Wsheet.cell(row,12).value).strip()
                qQualMin=str(Wsheet.cell(row,13).value).strip()
                qRxLevMin=str(Wsheet.cell(row,14).value).strip()
                hsdschSupport=str(Wsheet.cell(row,15).value).strip()
                edchSupport=str(Wsheet.cell(row,16).value).strip()
                edchTti2Support=str(Wsheet.cell(row,17).value).strip()
                enhancedL2Support=str(Wsheet.cell(row,18).value).strip()
                with open(os.path.abspath(os.path.join(output,"%s.mo"%RNCName)),'a') as f:
                     f.write(lteericssondt.crexternalutrancell(RNCName=RNCName,ExternalCell=ExternalCell,IurLinkID=IurLinkID,CID=CID,LAC=LAC,RAC=RAC,PSC=PSC,uarfcnDl=uarfcnDl,uarfcnUl=uarfcnUl,agpsEnabled=agpsEnabled,individualOffset=individualOffset,maxTxPowerUl=maxTxPowerUl,qQualMin=qQualMin,qRxLevMin=qRxLevMin,hsdschSupport=hsdschSupport,edchSupport=edchSupport,edchTti2Support=edchTti2Support,enhancedL2Support=enhancedL2Support))    

     def externalgsmcell(self):
         WK=openpyxl.load_workbook(self.workbook)
         Wsheet=WK.get_sheet_by_name("externalgsmcell")
         nrow=Wsheet.max_row
         if nrow>=2:
            dict={}
            output=os.path.abspath(os.path.join(os.getcwd(),"output_externalgsmcell_%s"%time.strftime("%y%m%d%H%M%S",time.localtime())))
            os.mkdir(output)
            for row in range(2,nrow+1):
                RNCName=str(Wsheet.cell(row,1).value).strip()
                ExternalGsmCell=str(Wsheet.cell(row,2).value).strip()
                ExternalGsmNetwork=str(Wsheet.cell(row,3).value).strip()
                cellIdentity=str(Wsheet.cell(row,4).value).strip()
                ncc=str(Wsheet.cell(row,5).value).strip()
                bcc=str(Wsheet.cell(row,6).value).strip()
                bcchFrequency=str(Wsheet.cell(row,7).value).strip()
                lac=str(Wsheet.cell(row,8).value).strip()
                bandIndicator=str(Wsheet.cell(row,9).value).strip()
                individualOffset="0" if Wsheet.cell(row,10).value is None else str(Wsheet.cell(row,10).value).strip()
                maxTxPowerUl="100" if Wsheet.cell(row,11).value is None else str(Wsheet.cell(row,11).value).strip()
                qRxLevMin="-109" if Wsheet.cell(row,12).value is None else str(Wsheet.cell(row,12).value).strip()
                with open(os.path.abspath(os.path.join(output,"%s.mo"%RNCName)),'a') as f:
                     f.write(lteericssondt.crexternalgsmcell(RNCName=RNCName,ExternalGsmCell=ExternalGsmCell,ExternalGsmNetwork=ExternalGsmNetwork,cellIdentity=cellIdentity,ncc=ncc,bcc=bcc,bcchFrequency=bcchFrequency,lac=lac,bandIndicator=bandIndicator,individualOffset=individualOffset,maxTxPowerUl=maxTxPowerUl,qRxLevMin=qRxLevMin))   





if __name__=='__main__':
     a=ltescriptsExcute(os.path.abspath(os.path.join(os.getcwd(),"DT_jiaoben_NEW.xlsx")))
     a.paramater()
     a.DUS52_3Grelation()
     a.DUS52_4Grelation()
     a.DUS52_2Grelation()
     a.MeasBasedCsfbTargetSelection()
     a.EUtranFreqRelation()
     a.changeFrequency()
     a.utranrelation()
     a.gsmneighbor()
     a.externalutrancell()
     a.externalgsmcell()
     print("The script output is in the output directory")


