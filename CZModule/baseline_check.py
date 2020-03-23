# -*- coding: utf8 -*-

import os,time
import csv
import numpy as np
import pandas as pd
import openpyxl


class baseline_check(object):

    def __init__(self,**kwargs):

        self.config=kwargs.get("baseline_config")
        self.kgetpaser_output=kwargs.get("kgetpaser_output")
        self.checkoutput_dir=os.path.abspath(os.path.join(os.getcwd(),"baseline_check_%s"%time.strftime("%y%m%d%H%M%S",time.localtime())))
        self.checkoutput_dir=kwargs.get("checkoutput_dir")

        if self.kgetpaser_output is None or self.config is None or self.checkoutput_dir is None:
            print ("Please enter kgetpaser_output and baseline file and checkoutput_dir parameters")

    def baseline_check(self):
        try:
                wkbook=openpyxl.load_workbook(self.config)
                wsheet=wkbook.get_sheet_by_name("baselineconfig")
                nrow=wsheet.max_row
                if nrow>=2:
                    checkoutput_file=os.path.abspath(os.path.join(self.checkoutput_dir,"baselinecheck.csv"))
                    header_dict={"MO":[]}
                    header_df=pd.DataFrame(header_dict)
                    header_df["parameter_value"]=""
                    header_df["parameter_name"]=""
                    header_df["baseline_value"]=""
                    header_df.to_csv(checkoutput_file,index=False)
                    for row in range(2,nrow+1):
                        parameters=wsheet.cell(row,3).value
                        baseline_value=wsheet.cell(row,4).value
                        MO=wsheet.cell(row,1).value
                        MO1=str(wsheet.cell(row,2).value)
                        if "%s.csv"%(MO.lower()) in [i.lower() for i in os.listdir(self.kgetpaser_output)]:
                            with open(os.path.abspath(os.path.join(self.kgetpaser_output,"%s.csv"%MO))) as f:
                                  df=pd.read_csv(f)

                                  if parameters.lower() in [j.lower() for j in df.columns]:
                                     if MO1=="None" or MO1.strip()=="":
                                        df=df[(df["" + parameters + ""]!=baseline_value)][["MO",parameters]]
                                        df["canshu_name"]=parameters
                                        df["baseline_value"]=baseline_value
                                        df.to_csv(checkoutput_file,mode='a',index=False,header=False)
                                     else:
                                         df=df[(df["" + parameters + ""]!=baseline_value) & (df["MO"].str.contains("%s=%s"%(MO,MO1)))][["MO",parameters]]
                                         df["canshu_name"]=parameters
                                         df["baseline_value"]=baseline_value
                                         df.to_csv(checkoutput_file,mode='a',index=False,header=False)





        except Exception as e:
               print("error",e)


        finally:
              wkbook.close()